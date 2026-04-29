"""
Premium eBook Designer - Custom PDF design for 20-page eBooks
"""

import logging
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY

logger = logging.getLogger(__name__)


class PremiumEbookDesigner:
    """Design and create premium eBooks"""
    
    def __init__(self, title: str, author: str, branding: dict = None):
        self.title = title
        self.author = author
        self.branding = branding or self._default_branding()
        self.max_pages = 20
    
    def _default_branding(self) -> dict:
        """Default branding colors"""
        return {
            'primary_color': '#2C3E50',
            'secondary_color': '#E74C3C',
            'accent_color': '#3498DB',
        }
    
    def create_ebook(self, content: dict, metadata: dict = None) -> str:
        """Create professional eBook PDF"""
        
        pdf_path = f"outputs/{self.title.replace(' ', '_')}.pdf"
        doc = SimpleDocTemplate(pdf_path, pagesize=letter)
        
        elements = []
        styles = self._create_styles()
        
        # Cover page
        elements.extend(self._create_cover_page(styles, metadata))
        elements.append(PageBreak())
        
        # Table of contents
        elements.extend(self._create_toc(content, styles))
        elements.append(PageBreak())
        
        # Content pages
        elements.extend(self._create_content_pages(content, styles))
        
        # Back matter
        elements.extend(self._create_back_matter(styles, metadata))
        
        # Build PDF
        try:
            doc.build(elements)
            logger.info(f"✅ eBook created: {pdf_path}")
            return pdf_path
        except Exception as e:
            logger.error(f"Error creating eBook: {str(e)}")
            return ""
    
    def _create_styles(self) -> dict:
        """Create custom styled styles"""
        
        styles = getSampleStyleSheet()
        
        primary_color = self.branding.get('primary_color', '#2C3E50')
        secondary_color = self.branding.get('secondary_color', '#E74C3C')
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=32,
            textColor=colors.HexColor(primary_color),
            spaceAfter=6,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold',
        )
        
        heading2_style = ParagraphStyle(
            'CustomHeading2',
            parent=styles['Heading2'],
            fontSize=18,
            textColor=colors.HexColor(secondary_color),
            spaceAfter=12,
            spaceBefore=12,
            fontName='Helvetica-Bold',
        )
        
        body_style = ParagraphStyle(
            'CustomBody',
            parent=styles['BodyText'],
            fontSize=11,
            leading=14,
            alignment=TA_JUSTIFY,
            textColor=colors.HexColor('#333333'),
        )
        
        return {
            'title': title_style,
            'heading2': heading2_style,
            'body': body_style,
        }
    
    def _create_cover_page(self, styles: dict, metadata: dict = None) -> list:
        """Create professional cover page"""
        
        elements = []
        elements.append(Spacer(1, 2*inch))
        
        # Title
        elements.append(Paragraph(self.title, styles['title']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Subtitle
        if metadata and metadata.get('subtitle'):
            elements.append(Paragraph(metadata['subtitle'], styles['heading2']))
        
        elements.append(Spacer(1, 1.5*inch))
        
        # Author
        elements.append(Paragraph(f"By {self.author}", styles['body']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Date
        date_str = datetime.now().strftime("%B %Y")
        elements.append(Paragraph(f"Published {date_str}", styles['body']))
        
        return elements
    
    def _create_toc(self, content: dict, styles: dict) -> list:
        """Create table of contents"""
        
        elements = []
        elements.append(Paragraph("Table of Contents", styles['heading2']))
        elements.append(Spacer(1, 0.2*inch))
        
        for i, section_title in enumerate(content.keys(), 1):
            elements.append(Paragraph(f"{i}. {section_title}", styles['body']))
            elements.append(Spacer(1, 0.05*inch))
        
        return elements
    
    def _create_content_pages(self, content: dict, styles: dict) -> list:
        """Create main content pages"""
        
        elements = []
        
        for section_title, section_content in content.items():
            elements.append(Paragraph(section_title, styles['heading2']))
            elements.append(Spacer(1, 0.1*inch))
            
            # Truncate if needed
            words = str(section_content).split()
            max_words = 500  # Approx 2.5 pages per section
            truncated = ' '.join(words[:max_words])
            
            elements.append(Paragraph(truncated, styles['body']))
            elements.append(Spacer(1, 0.3*inch))
        
        return elements
    
    def _create_back_matter(self, styles: dict, metadata: dict = None) -> list:
        """Create back matter (author bio, resources, etc.)"""
        
        elements = []
        elements.append(PageBreak())
        elements.append(Paragraph("About the Author", styles['heading2']))
        
        if metadata and metadata.get('author_bio'):
            elements.append(Paragraph(metadata['author_bio'], styles['body']))
        else:
            elements.append(Paragraph(
                "This premium digital product was created using advanced AI technology and market research.",
                styles['body']
            ))
        
        return elements


class TemplateGenerator:
    """Generate Excel templates"""
    
    def __init__(self):
        pass
    
    def create_template(self, template_type: str, purpose: str = "") -> str:
        """Create Excel template"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = template_type.replace('_', ' ').title()
            
            # Add headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            
            if template_type == "business_plan":
                headers = ["Section", "Description", "Status", "Notes"]
            elif template_type == "tracking":
                headers = ["Date", "Metric", "Value", "Target", "Progress"]
            else:
                headers = ["Item 1", "Item 2", "Item 3", "Item 4"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Save
            filepath = f"outputs/{template_type}_template.xlsx"
            wb.save(filepath)
            logger.info(f"✅ Template created: {filepath}")
            return filepath
        
        except ImportError:
            logger.error("openpyxl not installed")
            return ""
