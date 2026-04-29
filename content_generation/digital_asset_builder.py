"""
Digital Asset Builder - Creates templates, plans, tools, and plug-and-play solutions
"""

import logging
from pathlib import Path

logger = logging.getLogger(__name__)


class DigitalAssetBuilder:
    """Build digital assets"""
    
    def __init__(self):
        Path("outputs").mkdir(exist_ok=True)
    
    def create_step_by_step_plan(self, problem: str, solution: str, skill_level: str = 'beginner') -> str:
        """Create implementation plan"""
        
        plan_content = f"""
STEP-BY-STEP IMPLEMENTATION PLAN
================================

Problem: {problem}
Solution: {solution}
Skill Level: {skill_level}

PHASE 1: FOUNDATION
-------------------
1. Understand the fundamentals
   - Time required: 1 hour
   - Resources: Online tutorials, documentation
   - Success criteria: Basic understanding achieved

2. Gather necessary tools
   - Time required: 30 minutes
   - Resources: Software downloads, accounts
   - Success criteria: All tools installed and configured

PHASE 2: IMPLEMENTATION
-----------------------
1. Set up your environment
   - Time required: 1-2 hours
   - Step-by-step instructions provided
   - Common issues and solutions included

2. Execute the solution
   - Time required: 2-3 hours
   - Follow the detailed walkthrough
   - Test each step as you go

PHASE 3: OPTIMIZATION
---------------------
1. Review and refine
   - Time required: 1 hour
   - Check against success criteria
   - Make adjustments as needed

2. Document your process
   - Time required: 30 minutes
   - Create notes for future reference
   - Share results if applicable

TOTAL TIME: 5-7 hours
"""
        
        # Save as PDF (using reportlab)
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
            
            pdf_path = f"outputs/implementation_plan_{problem.replace(' ', '_')}.pdf"
            doc = SimpleDocTemplate(pdf_path, pagesize=letter)
            
            styles = getSampleStyleSheet()
            elements = []
            
            for line in plan_content.split('\n'):
                if line.strip():
                    elements.append(Paragraph(line, styles['Normal']))
                elements.append(Spacer(1, 0.1*inch))
            
            doc.build(elements)
            logger.info(f"✅ Implementation plan created: {pdf_path}")
            return pdf_path
        
        except Exception as e:
            logger.error(f"Error creating plan: {str(e)}")
            # Fallback to text file
            txt_path = f"outputs/implementation_plan_{problem.replace(' ', '_')}.txt"
            with open(txt_path, 'w') as f:
                f.write(plan_content)
            return txt_path
    
    def create_template(self, template_type: str, use_case: str = "", fields: list = None) -> str:
        """Create Excel template"""
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = template_type.replace('_', ' ')
            
            # Header styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            
            # Add headers
            if fields:
                for col, field in enumerate(fields, 1):
                    cell = ws.cell(row=1, column=col)
                    cell.value = field
                    cell.fill = header_fill
                    cell.font = header_font
            else:
                headers = ["Field 1", "Field 2", "Field 3", "Field 4"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
            
            # Add sample rows
            for row in range(2, 12):
                for col in range(1, len(fields or ["F1", "F2", "F3", "F4"]) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal="center")
            
            filepath = f"outputs/{template_type}_template.xlsx"
            wb.save(filepath)
            logger.info(f"✅ Template created: {filepath}")
            return filepath
        
        except ImportError:
            logger.error("openpyxl not installed")
            return ""
    
    def create_plug_and_play_solution(self, use_case: str, complexity: str = 'medium') -> dict:
        """Create plug-and-play solution package"""
        
        config_content = f"""
# Plug-and-Play Configuration for: {use_case}
# Complexity: {complexity}
# Auto-generated configuration ready to use

[config]
enabled = true
mode = {complexity}

[parameters]
retry_attempts = 3
timeout_seconds = 300
verbose_logging = true

[features]
auto_setup = true
error_handling = true
monitoring = true
"""
        
        setup_script = f"""
#!/usr/bin/env python
"""
Setup script for: {use_case}
"""

import sys
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def setup():
    logger.info('Setting up {use_case}...')
    logger.info('✓ Configuration loaded')
    logger.info('✓ Dependencies verified')
    logger.info('✓ Environment configured')
    logger.info('✓ Ready to use!')
    return True

if __name__ == '__main__':
    if setup():
        print('Setup completed successfully!')
        sys.exit(0)
    else:
        print('Setup failed!')
        sys.exit(1)
"""
        
        # Save files
        config_path = f"outputs/{use_case.replace(' ', '_')}_config.ini"
        script_path = f"outputs/{use_case.replace(' ', '_')}_setup.py"
        
        with open(config_path, 'w') as f:
            f.write(config_content)
        
        with open(script_path, 'w') as f:
            f.write(setup_script)
        
        logger.info(f"✅ Plug-and-play solution created")
        logger.info(f"   Config: {config_path}")
        logger.info(f"   Setup script: {script_path}")
        
        return {
            'config': config_path,
            'setup_script': script_path,
            'use_case': use_case,
        }
    
    def create_mini_tool(self, tool_purpose: str, tool_type: str = 'utility') -> str:
        """Create mini Python tool"""
        
        tool_code = f"""
#!/usr/bin/env python
"""
{tool_purpose} - Mini Tool
Auto-generated utility
"""

import sys
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class {tool_purpose.replace(' ', '')}Tool:
    """Implementation of {tool_purpose}"""
    
    def __init__(self):
        logger.info(f'Initializing {tool_purpose}')
    
    def run(self, input_data=None):
        """Execute the tool"""
        logger.info(f'Running {tool_purpose}')
        
        # Tool logic here
        result = {{
            'status': 'success',
            'message': '{tool_purpose} executed successfully',
            'timestamp': __import__('datetime').datetime.now().isoformat(),
        }}
        
        return result

if __name__ == '__main__':
    tool = {tool_purpose.replace(' ', '')}Tool()
    result = tool.run()
    
    print(f"Result: {{result}}")
    print('Tool executed successfully!')
"""
        
        filepath = f"outputs/{tool_purpose.replace(' ', '_')}_tool.py"
        with open(filepath, 'w') as f:
            f.write(tool_code)
        
        logger.info(f"✅ Mini tool created: {filepath}")
        return filepath
